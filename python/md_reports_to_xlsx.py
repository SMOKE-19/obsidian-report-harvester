from pathlib import Path
from datetime import datetime
import re

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
SCHEMA_PATH = ROOT / "report_schema.yaml"
OUTPUT_PATH = ROOT / "report_table.xlsx"
CREATED_AT_COLUMN = "파일 생성 시각"
FILENAME_COLUMN = "파일명"


def strip_link_markup(text: str, preserve_links: bool = False) -> str:
    text = re.sub(r"!\[\[[^\]]+\]\]", "", text)
    text = re.sub(r"!\[[^\]]*\]\([^)]+\)", "", text)

    if preserve_links:
        text = re.sub(r"\[\[[^\]]+\]\]", "", text)
        text = re.sub(r"\[[^\]]+\]\(([^)]+)\)", r"\1", text)
        return text

    text = re.sub(r"\[\[[^\]]+\]\]", "", text)
    text = re.sub(r"\[[^\]]+\]\([^)]+\)", "", text)
    text = re.sub(r"https?://\S+", "", text)
    return text


def clean_basic_text(text: str) -> str:
    text = text.strip()
    if text.startswith("==") and text.endswith("=="):
        text = text[2:-2].strip()
    return text.replace("`", "").strip()


def clean_text(text: str, preserve_links: bool = False) -> str:
    text = clean_basic_text(text)
    text = strip_link_markup(text, preserve_links)
    return re.sub(r"[ \t]+", " ", text).strip()


def strip_comment_blocks(text: str) -> str:
    return re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)


def split_sections(text: str) -> dict[str, list[str]]:
    sections: dict[str, list[str]] = {}
    current = None

    for raw_line in strip_comment_blocks(text).splitlines():
        line = raw_line.rstrip()
        heading = re.match(r"^###\s+(.+?)\s*$", line)
        if heading:
            current = heading.group(1).strip()
            sections[current] = []
            continue
        if current is not None:
            sections[current].append(line)

    return sections


def normalize_bullet(line: str, preserve_links: bool | None = False) -> tuple[int, str] | None:
    match = re.match(r"^(\s*)-\s+(.*)$", line)
    if not match:
        return None
    indent = len(match.group(1))
    if preserve_links is None:
        return indent, clean_basic_text(match.group(2))
    return indent, clean_text(match.group(2), preserve_links)


def format_nested_lines(lines: list[str], base_indent: int, preserve_links: bool = False) -> str:
    bullets = [normalize_bullet(line, preserve_links) for line in lines]
    bullets = [bullet for bullet in bullets if bullet is not None]
    if not bullets:
        return ""

    first_content_indent = min(indent for indent, _ in bullets)
    values = []
    for indent, value in bullets:
        if not value:
            continue
        level = max((indent - first_content_indent) // 2, 0)
        values.append(("  " * level) + value)
    return "\n".join(values).strip()


def parse_keyed_section(lines: list[str], key_specs: dict[str, dict]) -> dict[str, str]:
    result = {spec["column"]: "" for spec in key_specs.values()}
    current_key = None
    current_column = None
    current_indent = 0
    nested: list[str] = []

    def should_preserve_links(spec: dict) -> bool:
        column = spec.get("column", "")
        key = spec.get("key", "")
        return "link" in column.lower() or "link" in key.lower() or "링크" in column or "링크" in key

    def flush() -> None:
        nonlocal nested, current_key, current_column, current_indent
        if current_key and current_column and nested and key_specs[current_key].get("include_children"):
            existing = result.get(current_column, "")
            nested_text = format_nested_lines(
                nested,
                current_indent,
                preserve_links=should_preserve_links(key_specs[current_key]),
            )
            result[current_column] = "\n".join(part for part in [existing, nested_text] if part).strip()
        nested = []

    for line in lines:
        bullet = normalize_bullet(line, preserve_links=None)
        if bullet is None:
            continue

        indent, value = bullet
        key_match = re.match(r"^([^:]+):\s*(.*)$", value)
        if indent == 0 and key_match:
            flush()
            key = key_match.group(1).strip()
            if key in key_specs:
                current_key = key
                current_column = key_specs[key]["column"]
                current_indent = indent
                result[current_column] = clean_text(
                    key_match.group(2),
                    preserve_links=should_preserve_links(key_specs[key]),
                )
            else:
                current_key = None
                current_column = None
            continue

        if current_key:
            nested.append(line)

    flush()
    return result


def parse_section_body(lines: list[str], preserve_links: bool = False) -> str:
    return format_nested_lines(lines, 0, preserve_links)


def get_created_at(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_ctime).strftime("%Y-%m-%d %H:%M:%S")


def load_export_specs() -> list[dict]:
    with SCHEMA_PATH.open(encoding="utf-8") as file:
        schema = yaml.safe_load(file)

    report_schema = next(iter(schema.values()))
    if "headers" in report_schema:
        specs = []
        for header_spec in report_schema["headers"]:
            header = header_spec["header"]
            for export_spec in header_spec.get("export", []):
                spec = {"header": header, **export_spec}
                specs.append(spec)
        return specs

    specs = []
    metadata = report_schema["metadata"]
    for key in metadata["primary_keys"]:
        specs.append({"header": metadata["header"], "column": key, "key": key, "include_children": True})
    for section in report_schema["sections"]:
        specs.append({"header": section, "column": section, "content": "all"})
    return specs


def build_rows(markdown_paths: list[Path], export_specs: list[dict]) -> list[dict[str, str]]:
    rows = []
    for path in markdown_paths:
        text = path.read_text(encoding="utf-8")
        sections = split_sections(text)
        row = {
            CREATED_AT_COLUMN: get_created_at(path),
            FILENAME_COLUMN: path.name,
        }
        keyed_specs_by_header: dict[str, dict[str, dict]] = {}

        for spec in export_specs:
            if spec.get("key"):
                keyed_specs_by_header.setdefault(spec["header"], {})[spec["key"]] = spec
            elif spec.get("content") == "all":
                preserve_links = "link" in spec["column"].lower() or "링크" in spec["column"]
                row[spec["column"]] = parse_section_body(sections.get(spec["header"], []), preserve_links)

        for header, key_specs in keyed_specs_by_header.items():
            row.update(parse_keyed_section(sections.get(header, []), key_specs))

        rows.append(row)
    return rows


def write_xlsx(rows: list[dict[str, str]], columns: list[str]) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "reports"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, column_name in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row.get(column_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, column_name in enumerate(columns, start=1):
        max_length = len(column_name)
        for row_index in range(2, len(rows) + 2):
            value = sheet.cell(row=row_index, column=column_index).value or ""
            max_line = max((len(line) for line in str(value).splitlines()), default=0)
            max_length = max(max_length, max_line)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 45)

    for row_index in range(2, len(rows) + 2):
        max_lines = 1
        for column_index in range(1, len(columns) + 1):
            value = sheet.cell(row=row_index, column=column_index).value or ""
            max_lines = max(max_lines, len(str(value).splitlines()))
        sheet.row_dimensions[row_index].height = min(max(18, max_lines * 16), 160)

    try:
        workbook.save(OUTPUT_PATH)
        return OUTPUT_PATH
    except PermissionError:
        fallback_path = OUTPUT_PATH.with_name(f"{OUTPUT_PATH.stem}_clean{OUTPUT_PATH.suffix}")
        workbook.save(fallback_path)
        return fallback_path


def main() -> None:
    export_specs = load_export_specs()
    markdown_paths = sorted(
        path
        for path in ROOT.glob("*.md")
        if not path.name.lower().startswith("readme")
    )
    rows = build_rows(markdown_paths, export_specs)
    columns = [CREATED_AT_COLUMN, FILENAME_COLUMN, *[spec["column"] for spec in export_specs]]
    output_path = write_xlsx(rows, columns)
    print(f"created: {output_path}")
    print(f"rows: {len(rows)}")


if __name__ == "__main__":
    main()
